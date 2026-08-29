import csv
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

PERSIAN_TRANS = str.maketrans({
    'ي': 'ی', 'ى': 'ی', 'ئ': 'ی', 'ك': 'ک', 'ة': 'ه', 'ۀ': 'ه', 'ؤ': 'و',
    'إ': 'ا', 'أ': 'ا', 'ٱ': 'ا', 'آ': 'ا',
})
DIGIT_TRANS = str.maketrans('۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩', '01234567890123456789')


def normalize_text(value, compact=False):
    text = unicodedata.normalize('NFKC', str(value or '')).translate(PERSIAN_TRANS)
    text = text.replace('\u200c', ' ').replace('\u200f', ' ').translate(DIGIT_TRANS)
    text = re.sub(r'[^0-9A-Za-zآ-ی\s]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip().lower()
    return text.replace(' ', '') if compact else text


def normalize_phone(value):
    text = str(value or '').translate(DIGIT_TRANS)
    digits = re.sub(r'\D', '', text)
    if digits.startswith('98') and len(digits) == 12:
        digits = '0' + digits[2:]
    return digits


def normalize_tracking(value):
    text = str(value or '').translate(DIGIT_TRANS).strip()
    if 'e+' in text.lower():
        raise ValueError('کد رهگیری به حالت Scientific Number تبدیل شده؛ ستون بارکد را Text کنید.')
    return re.sub(r'\D', '', text)


def _header_map(row):
    result = {}
    for idx, value in enumerate(row):
        key = normalize_text(value, compact=True)
        if not key:
            continue
        if ('بارکد' in key or 'رهگیری' in key) and 'code' not in result:
            result['code'] = idx
        elif any(x in key for x in ('نامگیرنده', 'نامونامخانوادگی', 'گیرنده')) and 'name' not in result:
            result['name'] = idx
        elif any(x in key for x in ('مقصد', 'شهر')) and 'city' not in result:
            result['city'] = idx
        elif 'آدرس' in key and 'address' not in result:
            result['address'] = idx
        elif any(x in key for x in ('موبایل', 'تلفن', 'شمارهتماس')) and 'phone' not in result:
            result['phone'] = idx
    return result


def _parse_rows(rows):
    rows = [list(r) for r in rows]
    if not rows:
        raise ValueError('فایل خالی است.')

    header_index = None
    columns = None
    for i, row in enumerate(rows[:30]):
        found = _header_map(row)
        if 'code' in found and 'name' in found:
            header_index = i
            columns = found
            break

    # Vesta/post exports historically use C/E/H/I. Keep that fallback for Cutella.
    if columns is None:
        columns = {'code': 2, 'city': 4, 'name': 7, 'address': 8}
        start = 0
    else:
        start = header_index + 1

    output = []
    for row_no, row in enumerate(rows[start:], start=start + 1):
        def value(key):
            idx = columns.get(key)
            return row[idx] if idx is not None and idx < len(row) else ''

        code = normalize_tracking(value('code'))
        if not code:
            continue
        if len(code) < 10:
            continue
        output.append({
            'row': row_no,
            'code': code,
            'name': str(value('name') or '').strip(),
            'city': str(value('city') or '').strip(),
            'address': str(value('address') or '').strip(),
            'phone': normalize_phone(value('phone')),
        })

    if not output:
        raise ValueError('هیچ کد رهگیری معتبری در فایل پیدا نشد.')
    return output


def parse_tracking_file(path):
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix in {'.xlsx', '.xlsm'}:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        return _parse_rows(ws.iter_rows(values_only=True))
    if suffix == '.csv':
        with path.open('r', encoding='utf-8-sig', newline='') as handle:
            return _parse_rows(csv.reader(handle))
    raise ValueError('فرمت فایل باید XLSX، XLSM یا CSV باشد.')
